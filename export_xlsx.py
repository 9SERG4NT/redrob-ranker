#!/usr/bin/env python3
"""Export the ranked top-100 as a formatted .xlsx workbook.

The competition submission format is CSV; this produces a human-readable Excel
version of the same ranking for review/sharing. It reads the canonical ranking from
submission.csv and (optionally) enriches each row with a few profile facts looked up
from the candidates file.

    python export_xlsx.py --csv ./submission.csv --candidates ./candidates.jsonl --out ./submission.xlsx
"""

from __future__ import annotations

import argparse
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from redrob_ranker.loader import iter_candidates


def _load_enrichment(candidates_path, wanted):
    """One streaming pass to pull profile facts for the wanted candidate_ids."""
    info = {}
    if not candidates_path:
        return info
    for c in iter_candidates(candidates_path):
        cid = c.get("candidate_id")
        if cid in wanted:
            p = c.get("profile", {}) or {}
            info[cid] = {
                "name": p.get("anonymized_name", ""),
                "title": p.get("current_title", ""),
                "company": p.get("current_company", ""),
                "yoe": p.get("years_of_experience", ""),
                "location": p.get("location", ""),
                "country": p.get("country", ""),
            }
            if len(info) == len(wanted):
                break
    return info


def export(csv_path, out_path, candidates_path=None):
    with open(csv_path, "r", encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))
    wanted = {r["candidate_id"] for r in rows}
    info = _load_enrichment(candidates_path, wanted)
    enriched = bool(info)

    wb = Workbook()
    ws = wb.active
    ws.title = "Top 100"

    if enriched:
        headers = ["rank", "candidate_id", "name", "current_title", "years_experience",
                   "current_company", "location", "country", "score", "reasoning"]
    else:
        headers = ["rank", "candidate_id", "score", "reasoning"]

    # Header styling.
    head_fill = PatternFill("solid", fgColor="C9603F")
    head_font = Font(bold=True, color="FFFFFF")
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(vertical="center")

    for r in rows:
        rank = int(r["rank"])
        score = float(r["score"])
        base = [rank, r["candidate_id"]]
        if enriched:
            e = info.get(r["candidate_id"], {})
            base += [e.get("name", ""), e.get("title", ""), e.get("yoe", ""),
                     e.get("company", ""), e.get("location", ""), e.get("country", "")]
        base += [score, r.get("reasoning", "")]
        ws.append(base)

    # Column widths + wrap the reasoning column.
    widths = ({"rank": 6, "candidate_id": 15, "name": 18, "current_title": 26,
               "years_experience": 15, "current_company": 18, "location": 22,
               "country": 12, "score": 10, "reasoning": 90})
    for j, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 14)
    reason_col = headers.index("reasoning") + 1
    for row in range(2, len(rows) + 2):
        ws.cell(row=row, column=reason_col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=headers.index("score") + 1).number_format = "0.000000"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    wb.save(out_path)
    print(f"[xlsx] wrote {out_path} ({len(rows)} rows, {'enriched' if enriched else 'basic'})")


def main():
    ap = argparse.ArgumentParser(description="Export ranked top-100 to .xlsx")
    ap.add_argument("--csv", default="./submission.csv")
    ap.add_argument("--candidates", default=None, help="Optional: enrich rows with profile facts")
    ap.add_argument("--out", default="./submission.xlsx")
    args = ap.parse_args()
    export(args.csv, args.out, args.candidates)


if __name__ == "__main__":
    main()
